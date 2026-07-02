from datetime import date
from gestor_guias.excel_processor import hoy_colombia
from pathlib import Path

import pandas as pd

from gestor_guias.devoluciones import generate_devoluciones_report
from gestor_guias.recaudo import generate_recaudo_report
from gestor_guias.relacion_ce_rr import generate_relacion_ce_rr_report
from gestor_guias.repository import GuiaRepository
from gestor_guias.reports import (
    build_cierre_breakdown,
    build_monthly_breakdown,
    filter_by_date,
    generate_entregadas_operador_excel,
    generate_monthly_operator_report,
    generate_operator_report,
    generate_salidas_operador_excel,
    normalize_dataframe,
)
from openpyxl import load_workbook


def build_dataframe(guia: str, operador: str, estado: str, valor: str, servicio: str = "") -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "PLANILLA": "1",
                "SERVICIO": servicio,
                "GUIA": guia,
                "UNID": "1",
                "TIPO DE SERVICIO": "PT",
                "DESTINATARIO": "Persona",
                "MUNICIPIO": "SAN GIL",
                "VALOR": valor,
                "OPERADOR": operador,
                "ESTADO": estado,
                "CAUSAL": "",
                "F_INGRESO": "2026-06-10 00:00:00",
                "F_ENTREGA": "",
            }
        ]
    )


def test_build_cierre_breakdown_suma_efectivo_de_varios_operadores(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "E", "")
    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.update_tracking_fields("200", "KEVIN", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-10", operador="OMAR", gestionadas=1, ro=0, n=0, d=0, e=1,
        recaudado=10_000, bancos=0, nequi=0, envia=0, efectivo=10_000,
    )
    repository.guardar_cierre(
        fecha="2026-06-10", operador="KEVIN", gestionadas=1, ro=0, n=0, d=0, e=1,
        recaudado=20_000, bancos=5_000, nequi=0, envia=0, efectivo=15_000,
        gastos=2_000,
    )

    dataframe = normalize_dataframe(repository.to_dataframe())
    resumen = build_cierre_breakdown(repository, dataframe, date(2026, 6, 10))

    assert int(resumen["EFECTIVO"].sum()) == 25_000
    assert int(resumen.set_index("OPERADOR").loc["KEVIN", "GASTOS"]) == 2_000


def test_build_cierre_breakdown_incluye_cantidad_de_billetes_contados(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-10", operador="OMAR", gestionadas=1, ro=0, n=0, d=0, e=1,
        recaudado=10_000, bancos=0, nequi=0, envia=0, efectivo=10_000,
        denominaciones={50_000: 0, 10_000: 1, 1_000: 0},
    )

    dataframe = normalize_dataframe(repository.to_dataframe())
    resumen = build_cierre_breakdown(repository, dataframe, date(2026, 6, 10))

    fila = resumen.set_index("OPERADOR").loc["OMAR"]
    assert int(fila["BILLETES 10.000"]) == 1
    assert int(fila["BILLETES 50.000"]) == 0


def test_build_cierre_breakdown_incluye_cierre_sin_guias_del_dia(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe_con_fecha("100", "", "", "10000", "2026-06-05 00:00:00"))
    repository.update_tracking_fields("100", "OMAR", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-10", operador="OMAR", gestionadas=0, ro=0, n=0, d=0, e=0,
        recaudado=0, bancos=0, nequi=0, envia=0, efectivo=7_161_650,
    )

    dataframe = filter_by_date(normalize_dataframe(repository.to_dataframe()), date(2026, 6, 10))
    resumen = build_cierre_breakdown(repository, dataframe, date(2026, 6, 10))

    assert int(resumen.set_index("OPERADOR").loc["OMAR", "EFECTIVO"]) == 7_161_650


def test_generate_operator_report_incluye_detalle_de_guias_entregadas(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe_con_fecha("100", "", "", "10000", "2026-06-10 00:00:00"))
    repository.update_tracking_fields("100", "OMAR", "E", "")
    repository.save_consolidated(build_dataframe_con_fecha("200", "", "", "20000", "2026-06-10 00:00:00"))
    repository.update_tracking_fields("200", "KEVIN", "R", "")

    output_path = generate_operator_report(repository, tmp_path / "output", date(2026, 6, 10))

    workbook = load_workbook(output_path)
    assert "GUIAS ENTREGADAS" in workbook.sheetnames
    assert "CIERRE" in workbook.sheetnames
    sheet = workbook["GUIAS ENTREGADAS"]
    guias = [sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)]
    assert guias == ["100"]


def test_generate_salidas_operador_excel_solo_incluye_guias_en_reparto(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "KEVIN", "R", "")

    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.update_tracking_fields("200", "KEVIN", "E", "")

    repository.save_consolidated(build_dataframe("300", "", "", "30000"))
    repository.update_tracking_fields("300", "OMAR", "R", "")

    output_path = generate_salidas_operador_excel(repository, tmp_path / "output", "KEVIN", date(2026, 6, 10))

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"

    worksheet = load_workbook(output_path)["SALIDAS"]
    guias = [worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1)]
    assert "100" in guias
    assert "200" not in guias
    assert "300" not in guias


def test_generate_salidas_operador_excel_no_depende_de_la_fecha_de_planilla(tmp_path: Path) -> None:
    # La guia puede haber llegado en una planilla de otro dia y salir hoy;
    # el informe debe encontrarla igual, sin filtrar por F_INGRESO.
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.asignar_salida(["100"], "KEVIN", "R")

    output_path = generate_salidas_operador_excel(repository, tmp_path / "output", "KEVIN", date(2026, 6, 23))

    assert output_path.exists()
    assert output_path.stat().st_size > 0


def test_guias_en_salida_respeta_el_orden_de_registro(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    for guia in ("300", "100", "200"):
        repository.save_consolidated(build_dataframe(guia, "", "", "10000"))

    repository.asignar_salida(["300", "100"], "KEVIN", "R")
    repository.asignar_salida(["200"], "KEVIN", "R")

    guias = repository.guias_en_salida("KEVIN", "R")

    assert [guia["guia"] for guia in guias] == ["300", "100", "200"]


def test_generate_salidas_operador_excel_sin_guias(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_salidas_operador_excel(repository, tmp_path / "output", "KEVIN", date(2026, 6, 10))

    assert output_path.exists()


def test_generate_entregadas_operador_excel_solo_incluye_guias_entregadas(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.asignar_salida(["100"], "KEVIN", "R")
    repository.update_tracking_fields("100", "KEVIN", "E", "")

    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.asignar_salida(["200"], "KEVIN", "R")

    repository.save_consolidated(build_dataframe("300", "", "", "30000"))
    repository.asignar_salida(["300"], "OMAR", "R")
    repository.update_tracking_fields("300", "OMAR", "E", "")

    output_path = generate_entregadas_operador_excel(repository, tmp_path / "output", "KEVIN", hoy_colombia())

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"

    worksheet = load_workbook(output_path)["ENTREGAS"]
    guias = [worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1)]
    assert "100" in guias
    assert "200" not in guias
    assert "300" not in guias


def test_generate_entregadas_operador_excel_incluye_hoja_de_cierre_con_denominaciones(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_entregadas_operador_excel(
        repository,
        tmp_path / "output",
        "KEVIN",
        hoy_colombia(),
        resumen={"gestionadas": 2, "recaudado": 30_000, "efectivo": 30_000, "efectivo_contado": 30_000},
        denominaciones={"20000": 1, "10000": 1},
    )

    worksheet = load_workbook(output_path)["CIERRE"]
    valores = [
        cell.value
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=3)
        for cell in row
        if cell.value is not None
    ]
    assert "CONTEO DE EFECTIVO EN CAJA" in valores
    assert 20000 in valores
    assert 10000 in valores
    assert 20000 * 1 in valores


def test_generate_entregadas_operador_excel_sin_guias(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_entregadas_operador_excel(repository, tmp_path / "output", "KEVIN", hoy_colombia())

    assert output_path.exists()


def test_generate_recaudo_report(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "E", "")

    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.update_tracking_fields("200", "OMAR", "R", "")

    output_path = generate_recaudo_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"


def test_generate_recaudo_report_incluye_bancos_nequi_envia_gastos_y_adelanto(tmp_path: Path) -> None:
    import openpyxl

    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-10", operador="OMAR", gestionadas=1, ro=0, n=0, d=0, e=1,
        recaudado=10_000, bancos=3_000, nequi=2_000, envia=1_000, efectivo=4_000,
        gastos=500, adelanto_salario=1_500,
    )

    output_path = generate_recaudo_report(repository, tmp_path / "output", date(2026, 6, 10))

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook["RECAUDO"]
    fila = next(fila for fila in sheet.iter_rows(min_row=5) if fila[0].value == "OMAR")

    assert fila[3].value == 3_000
    assert fila[4].value == 2_000
    assert fila[5].value == 1_000
    assert fila[6].value == 500
    assert fila[7].value == 1_500

    fila_gastos_dia = next(
        fila for fila in sheet.iter_rows(min_row=1, max_col=1) if fila[0].value == "Gastos del dia"
    )
    valor_gastos_dia = sheet.cell(row=fila_gastos_dia[0].row, column=2).value
    assert "G" in valor_gastos_dia


def test_generate_recaudo_report_no_data(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_recaudo_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()


def test_generate_relacion_ce_rr_report(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "15000", "RR"))
    repository.update_tracking_fields("100", "JOHAN ORTIZ", "E", "")

    repository.save_consolidated(build_dataframe("200", "", "", "19900", "CE"))
    repository.update_tracking_fields("200", "JOHAN ORTIZ", "E", "")

    repository.save_consolidated(build_dataframe("300", "", "", "20000", "PT"))
    repository.update_tracking_fields("300", "JOHAN ORTIZ", "E", "")

    repository.save_consolidated(build_dataframe("400", "", "", "30000", "RR"))
    repository.update_tracking_fields("400", "JOHAN ORTIZ", "R", "")

    output_path = generate_relacion_ce_rr_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert output_path.stat().st_size > 0


def test_generate_relacion_ce_rr_report_resta_link_envia(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "15000", "RR"))
    repository.update_tracking_fields("100", "JOHAN ORTIZ", "E", "")

    repository.save_consolidated(build_dataframe("200", "", "", "19900", "CE"))
    repository.update_tracking_fields("200", "JOHAN ORTIZ", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-10", operador="JOHAN ORTIZ", gestionadas=2, ro=0, n=0, d=0, e=2,
        recaudado=34_900, bancos=0, nequi=0, envia=10_000, efectivo=24_900,
    )

    output_path = generate_relacion_ce_rr_report(repository, tmp_path / "output", date(2026, 6, 10))

    workbook = load_workbook(output_path)
    sheet = workbook["RELACION CE Y RR"]

    fila_total = next(fila for fila in sheet.iter_rows() if fila[0].value == "TOTAL")
    assert sheet.cell(row=fila_total[0].row, column=4).value == 34_900

    fila_link = next(fila for fila in sheet.iter_rows() if fila[0].value == "(-) LINK ENVIA")
    assert sheet.cell(row=fila_link[0].row, column=4).value == 10_000

    fila_total_recaudar = next(fila for fila in sheet.iter_rows() if fila[0].value == "TOTAL A RECAUDAR")
    assert sheet.cell(row=fila_total_recaudar[0].row, column=4).value == 24_900


def test_generate_relacion_ce_rr_report_no_data(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_relacion_ce_rr_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()


def test_generate_devoluciones_report(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "D", "DIRECCION ERRADA")

    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.update_tracking_fields("200", "OMAR", "E", "")

    output_path = generate_devoluciones_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"

    dataframe = pd.read_excel(output_path, dtype=str)
    assert list(dataframe["GUIA"]) == ["100"]
    assert dataframe.iloc[0]["CAUSAL"] == "DIRECCION ERRADA"


def build_dataframe_con_fecha(guia: str, operador: str, estado: str, valor: str, fecha: str) -> pd.DataFrame:
    dataframe = build_dataframe(guia, operador, estado, valor)
    dataframe["F_INGRESO"] = fecha
    return dataframe


def test_build_monthly_breakdown_calcula_efectividad_por_operador(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe_con_fecha("100", "", "", "10000", "2026-06-05 00:00:00"))
    repository.update_tracking_fields("100", "KEVIN", "E", "")
    repository.save_consolidated(build_dataframe_con_fecha("200", "", "", "20000", "2026-06-15 00:00:00"))
    repository.update_tracking_fields("200", "KEVIN", "R", "")
    repository.save_consolidated(build_dataframe_con_fecha("300", "", "", "30000", "2026-07-01 00:00:00"))
    repository.update_tracking_fields("300", "KEVIN", "E", "")

    repository.guardar_cierre(
        fecha="2026-06-05", operador="KEVIN", gestionadas=1, ro=0, n=0, d=0, e=1,
        recaudado=10_000, bancos=0, nequi=0, envia=0, efectivo=10_000,
        gastos=2_000, adelanto_salario=5_000,
    )

    dataframe = normalize_dataframe(repository.to_dataframe())
    resumen = build_monthly_breakdown(repository, dataframe, 2026, 6)

    fila = resumen.set_index("OPERADOR").loc["KEVIN"]
    assert int(fila["GESTIONADAS"]) == 2
    assert int(fila["ENTREGADAS"]) == 1
    assert fila["EFECTIVIDAD"] == 50.0
    assert int(fila["GASTOS"]) == 2_000
    assert int(fila["ADELANTO_SALARIO"]) == 5_000


def test_generate_monthly_operator_report(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    repository.save_consolidated(build_dataframe_con_fecha("100", "", "", "10000", "2026-06-05 00:00:00"))
    repository.update_tracking_fields("100", "KEVIN", "E", "")

    output_path = generate_monthly_operator_report(repository, tmp_path / "output", 2026, 6)

    assert output_path.exists()
    assert output_path.suffix == ".xlsx"


def test_generate_devoluciones_report_no_data(tmp_path: Path) -> None:
    repository = GuiaRepository(tmp_path / "guias.db")

    output_path = generate_devoluciones_report(repository, tmp_path / "output", date(2026, 6, 10))

    assert output_path.exists()


def _preparar_entregadas_del_mes(tmp_path: Path) -> GuiaRepository:
    repository = GuiaRepository(tmp_path / "guias.db")
    repository.save_consolidated(build_dataframe("100", "", "", "10000"))
    repository.update_tracking_fields("100", "OMAR", "E", "")
    repository.save_consolidated(build_dataframe("200", "", "", "20000"))
    repository.update_tracking_fields("200", "OMAR", "E", "")
    repository.save_consolidated(build_dataframe("300", "", "", "30000"))
    repository.update_tracking_fields("300", "KEVIN", "E", "")
    # Una guia sin entregar no debe aparecer en el informe.
    repository.save_consolidated(build_dataframe("400", "KEVIN", "R", "5000"))
    return repository


def test_generate_cierre_mensual_entregadas_incluye_promedio_por_empleado(tmp_path: Path) -> None:
    from gestor_guias.reports import generate_cierre_mensual_entregadas_excel

    repository = _preparar_entregadas_del_mes(tmp_path)
    hoy = hoy_colombia()

    ruta = generate_cierre_mensual_entregadas_excel(repository, tmp_path, hoy.year, hoy.month)

    workbook = load_workbook(ruta)
    assert "RESUMEN" in workbook.sheetnames
    assert "ENTREGADAS" in workbook.sheetnames

    resumen = list(workbook["RESUMEN"].values)
    filas = {row[0]: row for row in resumen[1:]}
    assert filas["TOTAL"][1] == 3
    assert filas["TOTAL"][2] == 60_000
    assert filas["PROMEDIO POR EMPLEADO"][1] == 1.5
    assert filas["PROMEDIO POR EMPLEADO"][2] == 30_000

    entregadas = list(workbook["ENTREGADAS"].values)
    guias = {row[2] for row in entregadas[1:]}
    assert guias == {"100", "200", "300"}


def test_entregadas_mes_une_archivo_con_zona_de_trabajo(tmp_path: Path) -> None:
    from gestor_guias.reports import entregadas_mes_dataframe

    repository = _preparar_entregadas_del_mes(tmp_path)
    hoy = hoy_colombia()
    # Las dos primeras se archivan (cierre del dia); la tercera se entrega despues.
    repository.update_tracking_fields("300", "KEVIN", "R", "")
    assert repository.archivar_entregadas() == 2
    repository.update_tracking_fields("300", "KEVIN", "E", "")

    entregadas = entregadas_mes_dataframe(repository, hoy.year, hoy.month)

    assert sorted(entregadas["GUIA"]) == ["100", "200", "300"]
    # Un mes sin entregas no trae nada.
    assert entregadas_mes_dataframe(repository, 2020, 1).empty


def test_generate_cierre_mensual_entregadas_pdf(tmp_path: Path) -> None:
    from gestor_guias.reports import generate_cierre_mensual_entregadas_pdf

    repository = _preparar_entregadas_del_mes(tmp_path)
    hoy = hoy_colombia()

    ruta = generate_cierre_mensual_entregadas_pdf(repository, tmp_path, hoy.year, hoy.month)

    assert ruta.exists()
    assert ruta.suffix == ".pdf"
    assert ruta.stat().st_size > 0
