from pathlib import Path

import pandas as pd

from gestor_guias.excel_processor import normalize_header_date
from gestor_guias.exporter import display_date, export_dataframe, prepare_for_export


def test_display_date_formats_iso_dates() -> None:
    assert display_date("2026-06-11 00:00:00") == "11/06/2026"
    assert display_date("2026-06-11") == "11/06/2026"
    assert display_date("") == ""
    assert display_date("texto raro") == "texto raro"


def test_display_date_round_trip_with_import() -> None:
    # El consolidado exportado (DD/MM/YYYY) debe poder reimportarse al formato interno.
    assert normalize_header_date(display_date("2026-06-11 00:00:00")) == "2026-06-11 00:00:00"


def test_prepare_for_export_only_changes_fecha() -> None:
    dataframe = pd.DataFrame([{"GUIA": "100", "F_INGRESO": "2026-06-11 00:00:00"}])

    result = prepare_for_export(dataframe)

    assert result.loc[0, "F_INGRESO"] == "11/06/2026"
    assert result.loc[0, "GUIA"] == "100"
    # El dataframe original no se modifica.
    assert dataframe.loc[0, "F_INGRESO"] == "2026-06-11 00:00:00"


def test_export_dataframe_marca_columna_guia_como_texto(tmp_path: Path) -> None:
    from datetime import date
    from openpyxl import load_workbook

    dataframe = pd.DataFrame([{"GUIA": "0140158", "F_INGRESO": "2026-06-11 00:00:00"}])

    output_path = export_dataframe(dataframe, tmp_path, date(2026, 6, 11))

    workbook = load_workbook(output_path)
    worksheet = workbook["Hoja1"]
    columna_guia = next(
        cell.column for cell in worksheet[1] if str(cell.value).strip().upper() == "GUIA"
    )
    assert worksheet.cell(row=2, column=columna_guia).number_format == "@"
