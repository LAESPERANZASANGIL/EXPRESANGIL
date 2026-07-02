from __future__ import annotations

from datetime import date
from pathlib import Path
import re

from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

from .exporter import MONTHS_ES, display_date
from .repository import GuiaRepository


DETAIL_COLUMNS = [
    "PLANILLA",
    "SERVICIO",
    "GUIA",
    "UNID",
    "TIPO DE SERVICIO",
    "DESTINATARIO",
    "DIRECCION",
    "MUNICIPIO",
    "VALOR",
    "OPERADOR",
    "ESTADO",
    "CAUSAL",
    "F_INGRESO",
    "F_ENTREGA",
]

# Estado que indica que la guia fue entregada y su valor recaudado.
ESTADO_RECAUDO = "E"

# Estado que indica que la guia salio en reparto con el operador (igual a
# operadores.ESTADO_SALIDA; se repite aqui para evitar un import circular).
ESTADO_SALIDA = "R"

# Denominaciones de billetes que un operador puede contar al cerrar el dia
# (igual a operadores.DENOMINACIONES; se repite aqui para evitar un import circular).
DENOMINACIONES = (100_000, 50_000, 20_000, 10_000, 5_000, 2_000, 1_000, 500, 200, 100, 50)


def generate_reports(source_file: Path, output_dir: Path, target_date: date) -> Path:
    detail = pd.read_excel(source_file, dtype=str).fillna("")
    dataframe = normalize_dataframe(detail)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"informes {target_date.day:02d} {MONTHS_ES[target_date.month]}.xlsx"

    general = build_general_summary(dataframe)
    by_operator = build_operator_summary(dataframe)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        general.to_excel(writer, index=False, sheet_name="GENERAL")
        by_operator.to_excel(writer, index=False, sheet_name="POR OPERADOR")
        detail.to_excel(writer, index=False, sheet_name="DETALLE")

        for sheet_name in writer.sheets:
            apply_report_format(writer.sheets[sheet_name])

    return output_path


def normalize_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    dataframe = dataframe.copy()
    for column in dataframe.columns:
        dataframe[column] = dataframe[column].fillna("").astype(str).str.strip()

    dataframe["OPERADOR"] = dataframe["OPERADOR"].replace("", "SIN OPERADOR")
    dataframe["ESTADO"] = dataframe["ESTADO"].replace("", "SIN ESTADO")
    dataframe["CAUSAL"] = dataframe["CAUSAL"].replace("", "SIN CAUSAL")
    dataframe["VALOR_NUMERICO"] = dataframe["VALOR"].apply(value_to_number)
    dataframe["UNID_NUMERICA"] = pd.to_numeric(dataframe["UNID"], errors="coerce").fillna(0).astype(int)
    return dataframe


def build_general_summary(dataframe: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"CONCEPTO": "TOTAL GUIAS", "VALOR": len(dataframe)},
        {"CONCEPTO": "TOTAL UNIDADES", "VALOR": int(dataframe["UNID_NUMERICA"].sum())},
        {"CONCEPTO": "TOTAL VALOR", "VALOR": int(dataframe["VALOR_NUMERICO"].sum())},
        {"CONCEPTO": "", "VALOR": ""},
        {"CONCEPTO": "GUIAS POR ESTADO", "VALOR": ""},
    ]

    for estado, cantidad in dataframe.groupby("ESTADO").size().sort_values(ascending=False).items():
        rows.append({"CONCEPTO": estado, "VALOR": int(cantidad)})

    rows.append({"CONCEPTO": "", "VALOR": ""})
    rows.append({"CONCEPTO": "GUIAS POR MUNICIPIO", "VALOR": ""})
    for municipio, cantidad in dataframe.groupby("MUNICIPIO").size().sort_values(ascending=False).items():
        rows.append({"CONCEPTO": municipio, "VALOR": int(cantidad)})

    return pd.DataFrame(rows)


def build_operator_summary(dataframe: pd.DataFrame) -> pd.DataFrame:
    return build_breakdown(dataframe, "OPERADOR")


def build_breakdown(dataframe: pd.DataFrame, column: str) -> pd.DataFrame:
    return (
        dataframe.groupby(column, as_index=False)
        .agg(
            GUIAS=("GUIA", "count"),
            UNIDADES=("UNID_NUMERICA", "sum"),
            VALOR=("VALOR_NUMERICO", "sum"),
        )
        .sort_values("GUIAS", ascending=False)
    )


def filter_by_date(dataframe: pd.DataFrame, target_date: date) -> pd.DataFrame:
    if "F_INGRESO" not in dataframe.columns:
        return dataframe.iloc[0:0]

    prefix = target_date.isoformat()
    return dataframe[dataframe["F_INGRESO"].astype(str).str.startswith(prefix)]


def filter_by_month(dataframe: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    if "F_INGRESO" not in dataframe.columns:
        return dataframe.iloc[0:0]

    prefix = f"{year:04d}-{month:02d}"
    return dataframe[dataframe["F_INGRESO"].astype(str).str.startswith(prefix)]


def build_cierre_breakdown(
    repository: GuiaRepository, dataframe: pd.DataFrame, target_date: date | None, operador: str = ""
) -> pd.DataFrame:
    if operador:
        dataframe = dataframe[dataframe["OPERADOR"] == operador]

    fecha_texto = target_date.isoformat() if target_date else ""

    operadores_con_cierre = repository.operadores_con_cierre(fecha_texto) if fecha_texto else []
    if operador:
        operadores_con_cierre = [nombre for nombre in operadores_con_cierre if nombre == operador]

    operadores = sorted({*dataframe["OPERADOR"].dropna().unique(), *operadores_con_cierre})

    columnas_billetes = [f"BILLETES {denominacion:,}".replace(",", ".") for denominacion in DENOMINACIONES]

    if not operadores:
        return pd.DataFrame(
            columns=[
                "OPERADOR",
                "GESTIONADAS",
                "RO",
                "N",
                "D",
                "E",
                "UNIDADES ENTREGADAS",
                "RECAUDADO",
                "BANCOS",
                "NEQUI",
                "ENVIA",
                "GASTOS",
                "ADELANTO_SALARIO",
                "EFECTIVO",
                *columnas_billetes,
            ]
        )

    filas = []
    for nombre_operador in operadores:
        guias_operador = dataframe[dataframe["OPERADOR"] == nombre_operador]
        gestionadas = len(guias_operador)
        ro = int((guias_operador["ESTADO"].str.upper() == "RO").sum())
        n = int((guias_operador["ESTADO"].str.upper() == "N").sum())
        d = int((guias_operador["ESTADO"].str.upper() == "D").sum())
        entregadas = guias_operador[guias_operador["ESTADO"].str.upper() == ESTADO_RECAUDO]
        e = len(entregadas)
        unidades_entregadas = int(entregadas["UNID_NUMERICA"].sum())
        recaudado = int(entregadas["VALOR_NUMERICO"].sum())

        cierre = repository.obtener_cierre(fecha_texto, nombre_operador) if fecha_texto else None
        bancos = cierre["bancos"] if cierre else 0
        nequi = cierre["nequi"] if cierre else 0
        envia = cierre["envia"] if cierre else 0
        gastos = cierre["gastos"] if cierre else 0
        adelanto_salario = cierre["adelanto_salario"] if cierre else 0
        efectivo = (
            cierre["efectivo"] if cierre
            else recaudado - (bancos + nequi + envia + gastos + adelanto_salario)
        )
        denominaciones_contadas = cierre["denominaciones"] if cierre else {}

        fila = {
            "OPERADOR": nombre_operador,
            "GESTIONADAS": gestionadas,
            "RO": ro,
            "N": n,
            "D": d,
            "E": e,
            "UNIDADES ENTREGADAS": unidades_entregadas,
            "RECAUDADO": recaudado,
            "BANCOS": bancos,
            "NEQUI": nequi,
            "ENVIA": envia,
            "GASTOS": gastos,
            "ADELANTO_SALARIO": adelanto_salario,
            "EFECTIVO": efectivo,
        }
        for denominacion, columna in zip(DENOMINACIONES, columnas_billetes):
            fila[columna] = int(denominaciones_contadas.get(denominacion, 0))
        filas.append(fila)

    return pd.DataFrame(filas).sort_values("GESTIONADAS", ascending=False)


def build_monthly_breakdown(repository: GuiaRepository, dataframe: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    monthly = filter_by_month(dataframe, year, month)

    columns = ["OPERADOR", "GASTOS", "ADELANTO_SALARIO", "GESTIONADAS", "ENTREGADAS", "UNIDADES ENTREGADAS", "EFECTIVIDAD"]
    if monthly.empty:
        return pd.DataFrame(columns=columns)

    gastos_adelantos = repository.sumar_gastos_adelantos_mes(year, month)
    operadores = sorted(monthly["OPERADOR"].dropna().unique())

    filas = []
    for nombre_operador in operadores:
        guias_operador = monthly[monthly["OPERADOR"] == nombre_operador]
        gestionadas = len(guias_operador)
        filas_entregadas = guias_operador[guias_operador["ESTADO"].str.upper() == ESTADO_RECAUDO]
        entregadas = len(filas_entregadas)
        unidades_entregadas = int(filas_entregadas["UNID_NUMERICA"].sum())
        efectividad = round(entregadas / gestionadas * 100, 1) if gestionadas else 0.0
        extra = gastos_adelantos.get(nombre_operador, {})

        filas.append(
            {
                "OPERADOR": nombre_operador,
                "GASTOS": extra.get("gastos", 0),
                "ADELANTO_SALARIO": extra.get("adelanto_salario", 0),
                "GESTIONADAS": gestionadas,
                "ENTREGADAS": entregadas,
                "UNIDADES ENTREGADAS": unidades_entregadas,
                "EFECTIVIDAD": efectividad,
            }
        )

    return pd.DataFrame(filas, columns=columns).sort_values("GESTIONADAS", ascending=False)


def generate_monthly_operator_report(
    repository: GuiaRepository, output_dir: Path, year: int, month: int
) -> Path:
    dataframe = normalize_dataframe(repository.to_dataframe())

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"informe mensual por operador {MONTHS_ES[month]} {year}.xlsx"

    summary = build_monthly_breakdown(repository, dataframe, year, month)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="POR OPERADOR")
        for sheet_name in writer.sheets:
            apply_report_format(writer.sheets[sheet_name])

    return output_path


COLUMNAS_ENTREGADAS_MES = [
    "PLANILLA", "SERVICIO", "GUIA", "UNID", "DESTINATARIO", "DIRECCION",
    "MUNICIPIO", "VALOR", "OPERADOR", "CAUSAL", "F_INGRESO", "F_ENTREGA",
]


def entregadas_mes_dataframe(repository: GuiaRepository, year: int, month: int) -> pd.DataFrame:
    """Guias entregadas del mes (archivo + zona de trabajo) como DataFrame normalizado."""
    rows = repository.entregadas_mes(year, month)
    data = [
        {
            "PLANILLA": row["planilla"],
            "SERVICIO": row["servicio"],
            "GUIA": row["guia"],
            "UNID": row["unid"],
            "DESTINATARIO": row["destinatario"],
            "DIRECCION": row["direccion"],
            "MUNICIPIO": row["municipio"],
            "VALOR": row["valor"],
            "OPERADOR": row["operador"],
            "ESTADO": row["estado"],
            "CAUSAL": row["causal"],
            "F_INGRESO": row["fecha"],
            "F_ENTREGA": row["ingreso"],
        }
        for row in rows
    ]
    return normalize_dataframe(pd.DataFrame(data, columns=COLUMNAS_ENTREGADAS_MES + ["ESTADO"]))


def build_entregadas_mes_resumen(entregadas: pd.DataFrame) -> pd.DataFrame:
    """Estadistica por operador con TOTAL y PROMEDIO POR EMPLEADO."""
    if entregadas.empty:
        return pd.DataFrame(columns=["OPERADOR", "GUIAS ENTREGADAS", "UNIDADES ENTREGADAS", "VALOR RECAUDADO"])
    resumen = (
        entregadas.groupby("OPERADOR")
        .agg(**{
            "GUIAS ENTREGADAS": ("GUIA", "count"),
            "UNIDADES ENTREGADAS": ("UNID_NUMERICA", "sum"),
            "VALOR RECAUDADO": ("VALOR_NUMERICO", "sum"),
        })
        .reset_index()
        .sort_values("GUIAS ENTREGADAS", ascending=False)
    )
    total_guias = int(resumen["GUIAS ENTREGADAS"].sum())
    total_unidades = int(resumen["UNIDADES ENTREGADAS"].sum())
    total_valor = int(resumen["VALOR RECAUDADO"].sum())
    empleados = len(resumen)
    return pd.concat(
        [
            resumen,
            pd.DataFrame(
                [
                    {
                        "OPERADOR": "TOTAL",
                        "GUIAS ENTREGADAS": total_guias,
                        "UNIDADES ENTREGADAS": total_unidades,
                        "VALOR RECAUDADO": total_valor,
                    },
                    {
                        "OPERADOR": "PROMEDIO POR EMPLEADO",
                        "GUIAS ENTREGADAS": round(total_guias / empleados, 1),
                        "UNIDADES ENTREGADAS": round(total_unidades / empleados, 1),
                        "VALOR RECAUDADO": round(total_valor / empleados),
                    },
                ]
            ),
        ],
        ignore_index=True,
    )


def generate_cierre_mensual_entregadas_excel(
    repository: GuiaRepository, output_dir: Path, year: int, month: int
) -> Path:
    """Informe final del mes: detalle de todas las entregadas y estadistica por operador."""
    entregadas = entregadas_mes_dataframe(repository, year, month)
    resumen = build_entregadas_mes_resumen(entregadas)
    if entregadas.empty:
        detalle = pd.DataFrame(columns=COLUMNAS_ENTREGADAS_MES)
    else:
        detalle = (
            entregadas[COLUMNAS_ENTREGADAS_MES]
            .assign(VALOR=entregadas["VALOR_NUMERICO"])
            .sort_values(["OPERADOR", "F_ENTREGA", "GUIA"])
            .reset_index(drop=True)
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"informe mensual entregadas {MONTHS_ES[month]} {year}.xlsx"
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="RESUMEN")
        detalle.to_excel(writer, index=False, sheet_name="ENTREGADAS")
        for sheet_name in writer.sheets:
            apply_report_format(writer.sheets[sheet_name])

    return output_path


def generate_cierre_mensual_entregadas_pdf(
    repository: GuiaRepository, output_dir: Path, year: int, month: int
) -> Path:
    """Version PDF del informe final del mes: estadistica por operador."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    entregadas = entregadas_mes_dataframe(repository, year, month)
    resumen = build_entregadas_mes_resumen(entregadas)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"informe mensual entregadas {MONTHS_ES[month]} {year}.pdf"
    )

    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"INFORME MENSUAL DE ENTREGADAS - {MONTHS_ES[month].upper()} {year}", estilos["Title"]),
        Paragraph("Oficina Expresangil - Estadistica por operador", estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    encabezado = ["OPERADOR", "GUIAS ENTREGADAS", "UNIDADES ENTREGADAS", "VALOR RECAUDADO"]
    filas = [encabezado]
    for _, row in resumen.iterrows():
        filas.append([
            str(row["OPERADOR"]),
            str(row["GUIAS ENTREGADAS"]),
            str(row["UNIDADES ENTREGADAS"]),
            f"$ {int(row['VALOR RECAUDADO']):,}".replace(",", "."),
        ])
    if len(filas) == 1:
        filas.append(["Sin entregas registradas en el mes", "", "", ""])

    tabla = Table(filas, colWidths=[7 * cm, 3.5 * cm, 3.5 * cm, 4 * cm])
    estilo_tabla = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]
    # Resalta las filas TOTAL y PROMEDIO POR EMPLEADO.
    for indice, fila in enumerate(filas):
        if fila[0] in ("TOTAL", "PROMEDIO POR EMPLEADO"):
            estilo_tabla.append(("BACKGROUND", (0, indice), (-1, indice), colors.HexColor("#FFF3CD")))
            estilo_tabla.append(("FONTNAME", (0, indice), (-1, indice), "Helvetica-Bold"))
    tabla.setStyle(TableStyle(estilo_tabla))
    elementos.append(tabla)

    documento = SimpleDocTemplate(str(output_path), pagesize=letter, title=output_path.stem)
    documento.build(elementos)

    return output_path


def build_movimiento_mensual(repository: GuiaRepository, year: int, month: int) -> pd.DataFrame:
    """Todas las guias gestionadas en el mes (por F_ENTREGA), incluido el archivo.

    Las entregadas (E) salen de entregadas_mes (archivo + zona de trabajo);
    las novedades (RO, N, D) siguen en la zona de trabajo y se toman de alli.
    """
    prefijo = f"{year:04d}-{month:02d}"
    dataframe = normalize_dataframe(repository.to_dataframe())
    gestionadas = dataframe[dataframe["F_ENTREGA"].astype(str).str.startswith(prefijo)]
    novedades = gestionadas[gestionadas["ESTADO"].str.strip().str.upper() != ESTADO_RECAUDO]
    entregadas = entregadas_mes_dataframe(repository, year, month)
    if novedades.empty:
        return entregadas
    if entregadas.empty:
        return novedades
    columnas = [col for col in entregadas.columns if col in novedades.columns]
    return pd.concat([novedades[columnas], entregadas[columnas]], ignore_index=True)


def build_rendimiento_mensual(repository: GuiaRepository, year: int, month: int) -> pd.DataFrame:
    """Rendimiento del mes por operador: gestionadas, entregadas, prestamos y promedio diario."""
    movimiento = build_movimiento_mensual(repository, year, month)
    columnas = [
        "OPERADOR", "GUIAS GESTIONADAS", "GUIAS ENTREGADAS",
        "GASTOS", "ADELANTO/PRESTAMO", "EFECTIVIDAD %", "PROMEDIO DEL MES",
    ]
    if movimiento.empty:
        return pd.DataFrame(columns=columnas)

    gastos_adelantos = repository.sumar_gastos_adelantos_mes(year, month)

    filas = []
    for nombre_operador in sorted(movimiento["OPERADOR"].dropna().unique()):
        guias_operador = movimiento[movimiento["OPERADOR"] == nombre_operador]
        gestionadas = len(guias_operador)
        filas_entregadas = guias_operador[guias_operador["ESTADO"].str.strip().str.upper() == ESTADO_RECAUDO]
        entregadas = len(filas_entregadas)
        # Promedio de entregas por dia trabajado (dias con alguna gestion en el mes).
        dias = guias_operador["F_ENTREGA"].astype(str).str[:10].nunique()
        promedio = round(entregadas / dias) if dias else 0
        efectividad = round(entregadas / gestionadas * 100, 1) if gestionadas else 0.0
        extra = gastos_adelantos.get(nombre_operador, {})
        filas.append(
            {
                "OPERADOR": nombre_operador,
                "GUIAS GESTIONADAS": gestionadas,
                "GUIAS ENTREGADAS": entregadas,
                "GASTOS": extra.get("gastos", 0),
                "ADELANTO/PRESTAMO": extra.get("adelanto_salario", 0),
                "EFECTIVIDAD %": efectividad,
                "PROMEDIO DEL MES": promedio,
            }
        )

    return pd.DataFrame(filas, columns=columnas).sort_values("GUIAS ENTREGADAS", ascending=False)


def _formato_pesos_pdf(valor: object) -> str:
    return f"$ {int(valor):,}".replace(",", ".")


def generate_rendimiento_mensual_operador_pdf(
    repository: GuiaRepository, output_dir: Path, operador: str, year: int, month: int
) -> Path:
    """Ficha PDF del movimiento mensual de UN operador (formato tipo tarjeta)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle

    rendimiento = build_rendimiento_mensual(repository, year, month)
    fila = rendimiento[rendimiento["OPERADOR"].str.upper() == operador.strip().upper()]
    datos = fila.iloc[0] if not fila.empty else None

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"informe mensual {operador} {MONTHS_ES[month]} {year}.pdf"
    )

    encabezado = [
        ["INFORME MENSUAL POR OPERADOR"],
        [operador.upper()],
        [f"{MONTHS_ES[month].upper()} {year}"],
    ]
    cuerpo = [
        ["GUIAS GESTIONADAS", str(int(datos["GUIAS GESTIONADAS"])) if datos is not None else "0"],
        ["GUIAS ENTREGADAS", str(int(datos["GUIAS ENTREGADAS"])) if datos is not None else "0"],
        ["GASTOS", _formato_pesos_pdf(datos["GASTOS"]) if datos is not None else "$ 0"],
        ["ADELANTO/PRESTAMO", _formato_pesos_pdf(datos["ADELANTO/PRESTAMO"]) if datos is not None else "$ 0"],
        ["EFECTIVIDAD", f"{datos['EFECTIVIDAD %']} %" if datos is not None else "0 %"],
        ["PROMEDIO DEL MES (ENTREGAS/DIA)", str(int(datos["PROMEDIO DEL MES"])) if datos is not None else "0"],
    ]

    tabla_titulo = Table(encabezado, colWidths=[14 * cm])
    tabla_titulo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (0, 0), 14),
        ("FONTSIZE", (0, 1), (0, 2), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1F3864")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    tabla_cuerpo = Table(cuerpo, colWidths=[9 * cm, 5 * cm])
    tabla_cuerpo.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))

    documento = SimpleDocTemplate(str(output_path), pagesize=letter, title=output_path.stem)
    documento.build([tabla_titulo, Spacer(1, 0.1 * cm), tabla_cuerpo])

    return output_path


def generate_rendimiento_mensual_pdf(
    repository: GuiaRepository, output_dir: Path, year: int, month: int
) -> Path:
    """PDF con el rendimiento mensual de TODOS los operadores."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rendimiento = build_rendimiento_mensual(repository, year, month)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / (
        f"informe mensual operadores {MONTHS_ES[month]} {year}.pdf"
    )

    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"INFORME MENSUAL DE OPERADORES - {MONTHS_ES[month].upper()} {year}", estilos["Title"]),
        Paragraph("Oficina Expresangil - Rendimiento mensual por operador", estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    filas = [[
        "OPERADOR", "GUIAS\nGESTIONADAS", "GUIAS\nENTREGADAS",
        "GASTOS", "ADELANTO/\nPRESTAMO", "EFECTIVIDAD", "PROMEDIO\nDEL MES",
    ]]
    for _, fila in rendimiento.iterrows():
        filas.append([
            str(fila["OPERADOR"]),
            str(int(fila["GUIAS GESTIONADAS"])),
            str(int(fila["GUIAS ENTREGADAS"])),
            _formato_pesos_pdf(fila["GASTOS"]),
            _formato_pesos_pdf(fila["ADELANTO/PRESTAMO"]),
            f"{fila['EFECTIVIDAD %']} %",
            str(int(fila["PROMEDIO DEL MES"])),
        ])
    if len(filas) == 1:
        filas.append(["Sin movimientos registrados en el mes", "", "", "", "", "", ""])

    tabla = Table(filas, colWidths=[6 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    documento = SimpleDocTemplate(
        str(output_path), pagesize=landscape(letter), title=output_path.stem
    )
    documento.build(elementos)

    return output_path


def generate_operator_report(
    repository: GuiaRepository, output_dir: Path, target_date: date | None = None, operador: str = ""
) -> Path:
    dataframe = normalize_dataframe(repository.to_dataframe())
    if target_date is not None:
        dataframe = filter_by_date(dataframe, target_date)

    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = f" {target_date.day:02d} {MONTHS_ES[target_date.month]}" if target_date else ""
    sufijo_operador = f" {operador}" if operador else ""
    output_path = output_dir / f"informe por operador{sufijo_operador}{suffix}.xlsx"

    summary = build_cierre_breakdown(repository, dataframe, target_date, operador)
    entregadas = build_entregadas_detalle(dataframe, operador)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="CIERRE")
        entregadas.to_excel(writer, index=False, sheet_name="GUIAS ENTREGADAS")
        for sheet_name in writer.sheets:
            apply_report_format(writer.sheets[sheet_name])

    return output_path


def build_entregadas_detalle(dataframe: pd.DataFrame, operador: str = "") -> pd.DataFrame:
    columnas = ["OPERADOR", "GUIA", "UNID", "DESTINATARIO", "DIRECCION", "MUNICIPIO", "VALOR", "F_ENTREGA"]
    entregadas = dataframe[dataframe["ESTADO"].str.upper() == ESTADO_RECAUDO]
    if operador:
        entregadas = entregadas[entregadas["OPERADOR"] == operador]

    if entregadas.empty:
        return pd.DataFrame(columns=columnas)

    return (
        entregadas[columnas]
        .assign(VALOR=entregadas["VALOR_NUMERICO"])
        .sort_values(["OPERADOR", "GUIA"])
        .reset_index(drop=True)
    )


# Meta diaria de guias entregadas por operador para calcular la efectividad.
META_DIARIA_GUIAS = 52


def generate_salidas_operador_excel(
    repository: GuiaRepository, output_dir: Path, operador: str, target_date: date
) -> Path:
    guias = repository.guias_en_salida(operador, ESTADO_SALIDA)

    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = f" {target_date.day:02d} {MONTHS_ES[target_date.month]}"
    output_path = output_dir / f"salidas {operador}{suffix}.xlsx"

    filas = []
    total_valor = 0
    total_unidades = 0
    for guia in guias:
        valor = value_to_number(guia.get("valor", ""))
        unidades = value_to_number(guia.get("unid", ""))
        total_valor += valor
        total_unidades += unidades
        filas.append(
            {
                "GUIA": guia.get("guia", ""),
                "UNID": unidades,
                "DESTINATARIO": guia.get("destinatario", ""),
                "DIRECCION": guia.get("direccion", ""),
                "VALOR": valor,
            }
        )

    filas.append({"GUIA": "", "UNID": total_unidades, "DESTINATARIO": "", "DIRECCION": "TOTAL", "VALOR": total_valor})

    dataframe = pd.DataFrame(filas, columns=["GUIA", "UNID", "DESTINATARIO", "DIRECCION", "VALOR"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="SALIDAS")
        apply_report_format(writer.sheets["SALIDAS"])

    return output_path


RESUMEN_CIERRE_ETIQUETAS = (
    ("gestionadas", "Guias gestionadas (salidas del dia)"),
    ("ro", "Reclama oficina (RO)"),
    ("n", "Novedades operativas (N)"),
    ("d", "Devoluciones (D)"),
    ("e", "Entregadas y recaudadas (E)"),
    ("unidades", "Unidades entregadas"),
    ("recaudado", "Dinero recaudado"),
    ("bancos", "Dinero en bancos"),
    ("nequi", "Dinero en Nequi"),
    ("envia", "Dinero en link Envia"),
    ("gastos", "Gastos"),
    ("adelanto_salario", "Adelanto de salario"),
    ("efectivo", "Efectivo a entregar"),
    ("efectivo_contado", "Efectivo contado en caja"),
    ("diferencia", "Diferencia"),
)
RESUMEN_CIERRE_CAMPOS_MONEDA = {
    "recaudado", "bancos", "nequi", "envia", "gastos", "adelanto_salario",
    "efectivo", "efectivo_contado", "diferencia",
}


def build_cierre_sheet(worksheet, resumen: dict, denominaciones: dict) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)

    worksheet.merge_cells("A1:B1")
    titulo = worksheet["A1"]
    titulo.value = "CIERRE DEL DIA"
    titulo.fill = header_fill
    titulo.font = header_font
    titulo.alignment = Alignment(horizontal="center")

    row = 2
    for clave, etiqueta in RESUMEN_CIERRE_ETIQUETAS:
        if clave not in resumen:
            continue
        worksheet.cell(row=row, column=1, value=etiqueta).font = label_font
        valor_celda = worksheet.cell(row=row, column=2, value=resumen[clave])
        if clave in RESUMEN_CIERRE_CAMPOS_MONEDA:
            valor_celda.number_format = '"$" #,##0'
        row += 1

    nota = str(resumen.get("nota") or "").strip()
    if nota:
        worksheet.cell(row=row, column=1, value="Anotacion").font = label_font
        worksheet.cell(row=row, column=2, value=nota)
        row += 1

    entradas = sorted(
        ((int(denominacion), int(cantidad)) for denominacion, cantidad in (denominaciones or {}).items() if int(cantidad) > 0),
        reverse=True,
    )
    if entradas:
        row += 1
        worksheet.merge_cells(f"A{row}:C{row}")
        cabecera = worksheet.cell(row=row, column=1, value="CONTEO DE EFECTIVO EN CAJA")
        cabecera.fill = header_fill
        cabecera.font = header_font
        cabecera.alignment = Alignment(horizontal="center")
        row += 1

        for columna, encabezado in enumerate(("Denominacion", "Cantidad", "Subtotal"), start=1):
            celda = worksheet.cell(row=row, column=columna, value=encabezado)
            celda.font = label_font
        row += 1

        for denominacion, cantidad in entradas:
            worksheet.cell(row=row, column=1, value=denominacion).number_format = '"$" #,##0'
            worksheet.cell(row=row, column=2, value=cantidad)
            subtotal_celda = worksheet.cell(row=row, column=3, value=denominacion * cantidad)
            subtotal_celda.number_format = '"$" #,##0'
            row += 1

    for columna, ancho in (("A", 32), ("B", 18), ("C", 18)):
        worksheet.column_dimensions[columna].width = ancho


def generate_entregadas_operador_excel(
    repository: GuiaRepository,
    output_dir: Path,
    operador: str,
    target_date: date,
    resumen: dict | None = None,
    denominaciones: dict | None = None,
) -> Path:
    guias = repository.guias_de_operador(operador, target_date.isoformat())
    entregadas = [
        guia for guia in guias if (guia.get("estado") or "").strip().upper() == ESTADO_RECAUDO
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = f" {target_date.day:02d} {MONTHS_ES[target_date.month]}"
    output_path = output_dir / f"entregas {operador}{suffix}.xlsx"

    filas = []
    total_valor = 0
    total_unidades = 0
    for guia in entregadas:
        valor = value_to_number(guia.get("valor", ""))
        unidades = value_to_number(guia.get("unid", ""))
        total_valor += valor
        total_unidades += unidades
        filas.append(
            {
                "GUIA": guia.get("guia", ""),
                "UNID": unidades,
                "DESTINATARIO": guia.get("destinatario", ""),
                "DIRECCION": guia.get("direccion", ""),
                "MUNICIPIO": guia.get("municipio", ""),
                "VALOR": valor,
            }
        )

    filas.append(
        {"GUIA": "", "UNID": total_unidades, "DESTINATARIO": "", "DIRECCION": "", "MUNICIPIO": "TOTAL", "VALOR": total_valor}
    )

    dataframe = pd.DataFrame(filas, columns=["GUIA", "UNID", "DESTINATARIO", "DIRECCION", "MUNICIPIO", "VALOR"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="ENTREGAS")
        apply_report_format(writer.sheets["ENTREGAS"])

        if resumen or denominaciones:
            cierre_sheet = writer.book.create_sheet("CIERRE")
            build_cierre_sheet(cierre_sheet, resumen or {}, denominaciones or {})

    return output_path


def generate_daily_report(repository: GuiaRepository, output_dir: Path, target_date: date) -> Path:
    dataframe = normalize_dataframe(repository.to_dataframe())
    daily = filter_by_date(dataframe, target_date)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"informe del dia {target_date.day:02d} {MONTHS_ES[target_date.month]}.xlsx"

    by_estado = build_breakdown(daily, "ESTADO")
    by_municipio = build_breakdown(daily, "MUNICIPIO")
    by_operador = build_breakdown(daily, "OPERADOR")
    detail = daily[DETAIL_COLUMNS]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        by_estado.to_excel(writer, index=False, sheet_name="POR ESTADO")
        by_municipio.to_excel(writer, index=False, sheet_name="POR MUNICIPIO")
        by_operador.to_excel(writer, index=False, sheet_name="POR OPERADOR")
        detail.to_excel(writer, index=False, sheet_name="DETALLE")

        for sheet_name in writer.sheets:
            apply_report_format(writer.sheets[sheet_name])

        resumen_sheet = writer.book.create_sheet("RESUMEN", 0)
        build_resumen_sheet(resumen_sheet, daily, target_date)

    return output_path


# Estilos del bloque RESUMEN del Informe Diario.
RESUMEN_DARK_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
RESUMEN_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
RESUMEN_HEADER_FONT = Font(bold=True, color="FFFFFF")
RESUMEN_LABEL_FONT = Font(bold=True)
RESUMEN_CURRENCY_FORMAT = '"$" #,##0.00'


def build_resumen_sheet(worksheet, daily: pd.DataFrame, target_date: date) -> None:
    fecha_label = f"{MONTHS_ES[target_date.month].upper()} {target_date.day} DE {target_date.year}"

    worksheet.merge_cells("A1:B1")
    title_cell = worksheet["A1"]
    title_cell.value = "INFORME DIARIO"
    title_cell.fill = RESUMEN_DARK_FILL
    title_cell.font = RESUMEN_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 22

    worksheet.merge_cells("A2:B2")
    fecha_cell = worksheet["A2"]
    fecha_cell.value = fecha_label
    fecha_cell.fill = RESUMEN_DARK_FILL
    fecha_cell.font = RESUMEN_HEADER_FONT
    fecha_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_unidades = int(daily["UNID_NUMERICA"].sum()) if not daily.empty else 0
    total_valor = int(daily["VALOR_NUMERICO"].sum()) if not daily.empty else 0

    row = 3
    for label, value, is_currency in (
        ("TOTAL GUIAS", len(daily), False),
        ("TOTAL UNIDADES", total_unidades, False),
        ("TOTAL VALOR", total_valor, True),
    ):
        label_cell = worksheet.cell(row=row, column=1, value=label)
        label_cell.font = RESUMEN_LABEL_FONT
        value_cell = worksheet.cell(row=row, column=2, value=value)
        value_cell.alignment = Alignment(horizontal="right")
        if is_currency:
            value_cell.number_format = RESUMEN_CURRENCY_FORMAT
        row += 1

    row += 1
    row = _write_resumen_breakdown(worksheet, row, "GUIAS POR ESTADO", daily, "ESTADO")
    row += 1
    row = _write_resumen_breakdown(worksheet, row, "GUIAS POR MUNICIPIO", daily, "MUNICIPIO")

    row += 1
    if daily.empty:
        recaudado = 0
    else:
        recaudado = int(
            daily[daily["ESTADO"].str.upper() == ESTADO_RECAUDO]["VALOR_NUMERICO"].sum()
        )

    label_cell = worksheet.cell(row=row, column=1, value=f"VALOR RECAUDADO ({ESTADO_RECAUDO}) $")
    label_cell.fill = RESUMEN_DARK_FILL
    label_cell.font = RESUMEN_HEADER_FONT
    label_cell.alignment = Alignment(horizontal="left", vertical="center")

    value_cell = worksheet.cell(row=row, column=2, value=recaudado)
    value_cell.fill = RESUMEN_DARK_FILL
    value_cell.font = RESUMEN_HEADER_FONT
    value_cell.alignment = Alignment(horizontal="right", vertical="center")
    value_cell.number_format = RESUMEN_CURRENCY_FORMAT

    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 18


def _write_resumen_breakdown(worksheet, row: int, title: str, daily: pd.DataFrame, column: str) -> int:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    title_cell = worksheet.cell(row=row, column=1, value=title)
    title_cell.fill = RESUMEN_DARK_FILL
    title_cell.font = RESUMEN_HEADER_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 1

    if daily.empty:
        return row

    for key, cantidad in daily.groupby(column).size().sort_values(ascending=False).items():
        worksheet.cell(row=row, column=1, value=key)
        value_cell = worksheet.cell(row=row, column=2, value=int(cantidad))
        value_cell.alignment = Alignment(horizontal="right")
        row += 1

    return row


# Planilla de Devoluciones/Entregadas: (encabezado de salida, columna origen).
# FECHA = fecha de entrega (F_ENTREGA); INGRESO = fecha de importacion (F_INGRESO).
GESTION_DETAIL_COLUMNS = [
    ("PLANILLA", "PLANILLA"),
    ("COBRO", "SERVICIO"),
    ("GUIA", "GUIA"),
    ("UNID", "UNID"),
    ("TIPO", "TIPO DE SERVICIO"),
    ("DESTINATARIO", "DESTINATARIO"),
    ("CIUDAD", "MUNICIPIO"),
    ("VALOR", "VALOR"),
    ("ESTADO", "ESTADO"),
    ("COD", "CAUSAL"),
    ("FECHA", "F_ENTREGA"),
    ("INGRESO", "F_INGRESO"),
]


def generate_estado_report(
    repository: GuiaRepository,
    output_dir: Path,
    target_date: date,
    estado: str,
    titulo: str,
) -> Path:
    """Planilla de toda la oficina con las guias de un estado (D o E), filtrando por
    la fecha de ENTREGA (F_ENTREGA). Una sola hoja con el formato de la operacion."""
    dataframe = repository.to_dataframe().fillna("").astype(str)
    prefijo = target_date.isoformat()
    seleccion = dataframe[
        (dataframe["ESTADO"].str.upper() == estado.strip().upper())
        & (dataframe["F_ENTREGA"].str.startswith(prefijo))
    ]

    detalle = pd.DataFrame()
    for encabezado, origen in GESTION_DETAIL_COLUMNS:
        columna = seleccion[origen]
        if origen in ("F_ENTREGA", "F_INGRESO"):
            columna = columna.map(display_date)
        detalle[encabezado] = columna.values

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{target_date.strftime('%d-%m-%Y')} - {titulo}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detalle.to_excel(writer, index=False, sheet_name=titulo[:31])
        apply_report_format(writer.sheets[titulo[:31]])

    return output_path


def generate_devoluciones_report(repository: GuiaRepository, output_dir: Path, target_date: date) -> Path:
    return generate_estado_report(repository, output_dir, target_date, "D", "Devoluciones")


def generate_entregadas_report(repository: GuiaRepository, output_dir: Path, target_date: date) -> Path:
    return generate_estado_report(repository, output_dir, target_date, ESTADO_RECAUDO, "Entregadas")


def value_to_number(value: object) -> int:
    text = str(value).strip()
    if not text or text == "$ -":
        return 0
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else 0


def apply_report_format(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    data_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_font = Font(bold=True)

    guia_columnas = {
        cell.column for cell in worksheet[1] if str(cell.value).strip().upper() == "GUIA"
    }

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.fill = data_fill
            if cell.column in guia_columnas:
                cell.number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 38)
