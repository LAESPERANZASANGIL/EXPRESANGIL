from __future__ import annotations

from datetime import date
from pathlib import Path
import re

from openpyxl.styles import Font, PatternFill
import pandas as pd


MONTHS_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


def output_filename(target_date: date) -> str:
    return f"{target_date.day:02d} {MONTHS_ES[target_date.month]}.xlsx"


def display_date(value: object) -> str:
    """Convierte la fecha guardada (YYYY-MM-DD [HH:MM:SS]) a DD/MM/YYYY para el Excel."""
    text = str(value).strip()
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        year, month, day = match.groups()
        return f"{day}/{month}/{year}"
    return text


def prepare_for_export(dataframe: pd.DataFrame) -> pd.DataFrame:
    result = dataframe.copy()
    for column in ("F_INGRESO", "F_ENTREGA"):
        if column in result.columns:
            result[column] = result[column].map(display_date)
    return result


def export_dataframe(dataframe: pd.DataFrame, output_dir: Path, target_date: date) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_filename(target_date)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        prepare_for_export(dataframe).to_excel(writer, index=False, sheet_name="Hoja1")
        worksheet = writer.sheets["Hoja1"]
        apply_master_format(worksheet)

    return output_path


def export_movements_copy(dataframe: pd.DataFrame, output_dir: Path, target_date: date) -> Path | None:
    if dataframe.empty:
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"movimientos otros estado {output_filename(target_date)}"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        prepare_for_export(dataframe).to_excel(writer, index=False, sheet_name="Movimientos")
        worksheet = writer.sheets["Movimientos"]
        apply_master_format(worksheet)

    return output_path


def apply_master_format(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    data_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    planilla_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    header_font = Font(bold=True, color="FFFFFF")
    planilla_font = Font(bold=True)

    widths = {
        "A": 13,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 18,
        "F": 35,
        "G": 13,
        "H": 13,
        "I": 13,
        "J": 13,
        "K": 13,
        "L": 20,
        "M": 13,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell.alignment.copy(horizontal="left")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            if cell.column == 1:
                cell.fill = planilla_fill
                cell.font = planilla_font
            else:
                cell.fill = data_fill
            cell.alignment = cell.alignment.copy(horizontal="left")
