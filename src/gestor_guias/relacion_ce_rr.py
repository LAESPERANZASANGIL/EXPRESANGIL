from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .exporter import MONTHS_ES
from .reports import ESTADO_RECAUDO, filter_by_date, normalize_dataframe
from .repository import GuiaRepository


OFICINA_NOMBRE = "SAN GIL"
ADMIN_NAME = "JOHAN A. ORTIZ"
SERVICIOS_RELACION = ("RR", "CE")

DARK_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
TITLE_FONT = Font(bold=True, color="FFC000")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def format_currency_co(value: int) -> str:
    return f"$ {value:,}".replace(",", ".")


def generate_relacion_ce_rr_report(
    repository: GuiaRepository,
    output_dir: Path,
    target_date: date,
    admin_name: str = ADMIN_NAME,
    oficina_nombre: str = OFICINA_NOMBRE,
) -> Path:
    dataframe = normalize_dataframe(repository.to_dataframe())
    daily = filter_by_date(dataframe, target_date)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"relacion guias ce y rr {target_date.day:02d} {MONTHS_ES[target_date.month]}.xlsx"

    fecha_label = target_date.strftime("%d/%m/%Y")

    if not daily.empty:
        relevant = daily[
            (daily["ESTADO"].str.upper() == ESTADO_RECAUDO) & (daily["SERVICIO"].str.upper().isin(SERVICIOS_RELACION))
        ].sort_values(["OPERADOR", "SERVICIO"])
    else:
        relevant = daily

    link_envia = repository.sumar_envia_dia(target_date.isoformat())

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "RELACION CE Y RR"
    _escribir_hoja_relacion(worksheet, relevant, admin_name, oficina_nombre, fecha_label, link_envia)

    workbook.save(output_path)

    return output_path


def _escribir_hoja_relacion(worksheet, rows, admin_name, oficina_nombre, fecha_label, link_envia: int = 0) -> None:
    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "RELACION DE GUIAS CE Y RR " + oficina_nombre
    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = "INFORME OFICINA EXPRESANGIL"
    worksheet["A3"] = admin_name
    worksheet["C3"] = "FECHA"
    worksheet["D3"] = fecha_label
    worksheet.merge_cells("A4:D4")
    worksheet["A4"] = "GUIAS CONTRAENTREGA Y RECAUDO"

    for fila in (1, 2, 4):
        for columna in range(1, 5):
            celda = worksheet.cell(row=fila, column=columna)
            celda.fill = DARK_FILL
            celda.font = TITLE_FONT
            celda.alignment = CENTER

    encabezados = ["N°", "CE O RR", "N° DE GUIA", "VALOR"]
    for columna, encabezado in enumerate(encabezados, start=1):
        celda = worksheet.cell(row=5, column=columna, value=encabezado)
        celda.fill = DARK_FILL
        celda.font = HEADER_FONT
        celda.alignment = CENTER

    fila_actual = 6
    total = 0
    for indice, (_, row) in enumerate(rows.iterrows(), start=1):
        valor = int(row["VALOR_NUMERICO"])
        total += valor
        worksheet.cell(row=fila_actual, column=1, value=indice).alignment = CENTER
        worksheet.cell(row=fila_actual, column=2, value=row["SERVICIO"]).alignment = CENTER
        guia_cell = worksheet.cell(row=fila_actual, column=3, value=row["GUIA"])
        guia_cell.number_format = "@"
        worksheet.cell(row=fila_actual, column=4, value=valor).number_format = '"$" #,##0'
        fila_actual += 1

    if fila_actual == 6:
        worksheet.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=4)
        worksheet.cell(row=fila_actual, column=1, value="Sin registros para esta fecha").alignment = CENTER
        fila_actual += 1

    worksheet.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=3)
    total_label_cell = worksheet.cell(row=fila_actual, column=1, value="TOTAL")
    total_label_cell.fill = DARK_FILL
    total_label_cell.font = HEADER_FONT
    total_label_cell.alignment = CENTER
    total_value_cell = worksheet.cell(row=fila_actual, column=4, value=total)
    total_value_cell.fill = DARK_FILL
    total_value_cell.font = HEADER_FONT
    total_value_cell.number_format = '"$" #,##0'
    fila_actual += 1

    worksheet.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=3)
    worksheet.cell(row=fila_actual, column=1, value="(-) LINK ENVIA").alignment = CENTER
    worksheet.cell(row=fila_actual, column=4, value=link_envia).number_format = '"$" #,##0'
    fila_actual += 1

    worksheet.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=3)
    total_recaudar_label_cell = worksheet.cell(row=fila_actual, column=1, value="TOTAL A RECAUDAR")
    total_recaudar_label_cell.fill = DARK_FILL
    total_recaudar_label_cell.font = HEADER_FONT
    total_recaudar_label_cell.alignment = CENTER
    total_recaudar_value_cell = worksheet.cell(row=fila_actual, column=4, value=total - link_envia)
    total_recaudar_value_cell.fill = DARK_FILL
    total_recaudar_value_cell.font = HEADER_FONT
    total_recaudar_value_cell.number_format = '"$" #,##0'

    worksheet.column_dimensions["A"].width = 6
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 14
