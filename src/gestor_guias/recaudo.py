from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .exporter import MONTHS_ES
from .reports import (
    DENOMINACIONES,
    DETAIL_COLUMNS,
    ESTADO_RECAUDO,
    apply_report_format,
    build_breakdown,
    build_resumen_sheet,
    filter_by_date,
    normalize_dataframe,
)
from .repository import GuiaRepository


CURRENCY_FORMAT = '"$" #,##0'

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
INPUT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="C6E0B4")
WARNING_FONT = Font(bold=True, color="C00000")
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))

EXTRA_OPERATOR_ROWS = 5

COLUMNS = ["OPERADOR", "UNID", "VALOR RECAUDADO", "BANCOS", "NEQUI", "ENVIA", "GASTOS", "ADELANTO SALARIO", "TOTAL"]


def generate_recaudo_report(repository: GuiaRepository, output_dir: Path, target_date: date) -> Path:
    """Genera el INFORME DIARIO: recaudo (con cierre general) + resumen del dia en un solo Excel."""
    dataframe = normalize_dataframe(repository.to_dataframe())
    daily = filter_by_date(dataframe, target_date)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"informe diario {target_date.day:02d} {MONTHS_ES[target_date.month]}.xlsx"

    if daily.empty:
        totals: dict[str, int] = {}
        recaudo: dict[str, float] = {}
        total_guias = 0
    else:
        totals = daily.groupby("OPERADOR")["GUIA"].count().to_dict()
        recaudo_series = (
            daily[daily["ESTADO"].str.upper() == ESTADO_RECAUDO].groupby("OPERADOR")["VALOR_NUMERICO"].sum()
        )
        recaudo = recaudo_series.to_dict()
        total_guias = len(daily)

    # Todos los operadores del dia: los que gestionaron guias Y los que
    # registraron cierre (un operador puede no tener entregas pero si
    # gastos o adelantos en su cierre).
    con_cierre = repository.operadores_con_cierre(target_date.isoformat())
    operators = sorted(
        {*totals, *con_cierre},
        key=lambda operador: (-totals.get(operador, 0), operador),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RECAUDO"

    last_column = len(COLUMNS)
    last_column_letter = get_column_letter(last_column)

    sheet.merge_cells(f"A1:{last_column_letter}1")
    title_cell = sheet["A1"]
    title_cell.value = "INFORME DE RECAUDO DIARIO"
    title_cell.fill = HEADER_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    sheet["A2"] = "FECHA:"
    sheet["A2"].font = Font(bold=True)
    sheet.merge_cells(f"B2:{last_column_letter}2")
    sheet["B2"] = target_date.isoformat()
    sheet["B2"].alignment = Alignment(horizontal="left")

    header_row = 4
    for column_index, column_name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    first_data_row = header_row + 1
    row_index = first_data_row

    for operador in operators:
        sheet.cell(row=row_index, column=1, value=operador).border = THIN_BORDER
        sheet.cell(row=row_index, column=2, value=int(totals.get(operador, 0))).border = THIN_BORDER

        valor_recaudado_cell = sheet.cell(row=row_index, column=3, value=int(recaudo.get(operador, 0)))
        valor_recaudado_cell.number_format = CURRENCY_FORMAT
        valor_recaudado_cell.border = THIN_BORDER

        cierre = repository.obtener_cierre(target_date.isoformat(), operador)
        valores_cierre = (
            (cierre.get("bancos", 0), cierre.get("nequi", 0), cierre.get("envia", 0),
             cierre.get("gastos", 0), cierre.get("adelanto_salario", 0))
            if cierre
            else (0, 0, 0, 0, 0)
        )
        for column_index, valor in zip((4, 5, 6, 7, 8), valores_cierre):
            cell = sheet.cell(row=row_index, column=column_index, value=int(valor or 0))
            cell.fill = INPUT_FILL
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

        total_cell = sheet.cell(
            row=row_index,
            column=9,
            value=f"=C{row_index}-D{row_index}-E{row_index}-F{row_index}-G{row_index}-H{row_index}",
        )
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.border = THIN_BORDER

        row_index += 1

    for _ in range(EXTRA_OPERATOR_ROWS):
        for column_index in range(1, last_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER
            if column_index in (3, 4, 5, 6, 7, 8):
                cell.number_format = CURRENCY_FORMAT
            if column_index in (4, 5, 6, 7, 8):
                cell.fill = INPUT_FILL

        total_cell = sheet.cell(
            row=row_index,
            column=9,
            value=f"=C{row_index}-D{row_index}-E{row_index}-F{row_index}-G{row_index}-H{row_index}",
        )
        total_cell.number_format = CURRENCY_FORMAT
        row_index += 1

    last_data_row = row_index - 1
    total_row = row_index

    total_label_cell = sheet.cell(row=total_row, column=1, value="TOTAL GENERAL")
    total_label_cell.fill = HEADER_FILL
    total_label_cell.font = HEADER_FONT

    for column_index in range(2, last_column + 1):
        column_letter = get_column_letter(column_index)
        cell = sheet.cell(
            row=total_row,
            column=column_index,
            value=f"=SUM({column_letter}{first_data_row}:{column_letter}{last_data_row})",
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        if column_index >= 3:
            cell.number_format = CURRENCY_FORMAT

    valor_recaudado_total_cell = f"C{total_row}"
    bancos_total_cell = f"D{total_row}"
    nequi_total_cell = f"E{total_row}"
    envia_total_cell = f"F{total_row}"
    gastos_total_cell = f"G{total_row}"
    adelantos_total_cell = f"H{total_row}"

    summary_row = total_row + 2
    sheet.merge_cells(f"A{summary_row}:B{summary_row}")
    summary_title_cell = sheet[f"A{summary_row}"]
    summary_title_cell.value = "RESUMEN DE VERIFICACION"
    summary_title_cell.fill = HEADER_FILL
    summary_title_cell.font = HEADER_FONT
    summary_title_cell.alignment = Alignment(horizontal="left", vertical="center")

    summary_rows = [
        ("Total guias del dia", total_guias, None),
        ("Valor total", float(daily["VALOR_NUMERICO"].sum()) if not daily.empty else 0, None),
        ("Valor total recaudado (= lista)", f"={valor_recaudado_total_cell}-{envia_total_cell}", HIGHLIGHT_FILL),
        ("Total a pagar (Bancos + Nequi)", f"={bancos_total_cell}+{nequi_total_cell}", None),
        ("Gastos del dia", f"={gastos_total_cell}", None),
        ("Adelantos del dia", f"={adelantos_total_cell}", None),
        (
            "Saldo pendiente (= recaudo - pagar - gastos - adelantos)",
            f"=B{summary_row + 3}-B{summary_row + 4}-B{summary_row + 5}-B{summary_row + 6}",
            None,
        ),
    ]

    for offset, (label, value, fill) in enumerate(summary_rows, start=1):
        label_cell = sheet.cell(row=summary_row + offset, column=1, value=label)
        value_cell = sheet.cell(row=summary_row + offset, column=2, value=value)
        value_cell.number_format = CURRENCY_FORMAT
        if fill:
            label_cell.fill = fill
            value_cell.fill = fill
        if "Saldo pendiente" in label:
            label_cell.font = WARNING_FONT
            value_cell.font = WARNING_FONT

    # --- Tabla del cierre general del dia (conteo de billetes de la oficina) ---
    cierre_general = repository.obtener_cierre_general(target_date.isoformat())
    denominaciones_contadas = cierre_general["denominaciones"] if cierre_general else {}

    cierre_title_row = summary_row + len(summary_rows) + 2
    sheet.merge_cells(f"A{cierre_title_row}:C{cierre_title_row}")
    cierre_title = sheet[f"A{cierre_title_row}"]
    cierre_title.value = "CIERRE GENERAL DEL DIA (EFECTIVO CONTADO)"
    cierre_title.fill = HEADER_FILL
    cierre_title.font = HEADER_FONT
    cierre_title.alignment = Alignment(horizontal="left", vertical="center")

    cierre_header_row = cierre_title_row + 1
    for column_index, column_name in enumerate(("DENOMINACION", "CANTIDAD", "SUBTOTAL"), start=1):
        cell = sheet.cell(row=cierre_header_row, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    fila_cierre = cierre_header_row + 1
    for denominacion in DENOMINACIONES:
        cantidad = int(denominaciones_contadas.get(denominacion, 0))
        denominacion_cell = sheet.cell(row=fila_cierre, column=1, value=denominacion)
        denominacion_cell.number_format = CURRENCY_FORMAT
        denominacion_cell.border = THIN_BORDER
        cantidad_cell = sheet.cell(row=fila_cierre, column=2, value=cantidad)
        cantidad_cell.fill = INPUT_FILL
        cantidad_cell.border = THIN_BORDER
        subtotal_cell = sheet.cell(row=fila_cierre, column=3, value=f"=A{fila_cierre}*B{fila_cierre}")
        subtotal_cell.number_format = CURRENCY_FORMAT
        subtotal_cell.border = THIN_BORDER
        fila_cierre += 1

    subtotal_row = fila_cierre
    subtotal_label = sheet.cell(row=subtotal_row, column=1, value="SUBTOTAL CONTADO")
    subtotal_label.fill = HEADER_FILL
    subtotal_label.font = HEADER_FONT
    sheet.cell(row=subtotal_row, column=2).fill = HEADER_FILL
    subtotal_total = sheet.cell(
        row=subtotal_row, column=3,
        value=f"=SUM(C{cierre_header_row + 1}:C{subtotal_row - 1})",
    )
    subtotal_total.fill = HEADER_FILL
    subtotal_total.font = HEADER_FONT
    subtotal_total.number_format = CURRENCY_FORMAT

    widths = {"A": 28, "B": 12, "C": 18, "D": 14, "E": 14, "F": 14, "G": 14, "H": 18, "I": 16}
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width

    # --- Hojas del informe del dia (antes eran un archivo aparte) ---
    resumen_sheet = workbook.create_sheet("RESUMEN")
    build_resumen_sheet(resumen_sheet, daily, target_date)

    # POR OPERADOR relaciona solo las guias entregadas (E) de cada operador.
    entregadas_dia = daily[daily["ESTADO"].str.upper() == ESTADO_RECAUDO]
    hojas_desglose = (
        ("POR ESTADO", build_breakdown(daily, "ESTADO")),
        ("POR MUNICIPIO", build_breakdown(daily, "MUNICIPIO")),
        ("POR OPERADOR", build_breakdown(entregadas_dia, "OPERADOR")),
        ("DETALLE", daily[DETAIL_COLUMNS] if not daily.empty else daily.reindex(columns=DETAIL_COLUMNS)),
    )
    for nombre_hoja, datos in hojas_desglose:
        hoja = workbook.create_sheet(nombre_hoja)
        hoja.append(list(datos.columns))
        for fila in datos.itertuples(index=False):
            hoja.append(list(fila))
        apply_report_format(hoja)

    workbook.save(output_path)
    return output_path
